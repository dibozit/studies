import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

#dataframe_to_rows > this will allow us to convert our data frame into a format that's usable by openpyxl

wb = load_workbook('regions.xlsx')
ws = wb.active

#burada pandada minik bi analiz yaptık
df = pd.read_excel('all_shifts.xlsx')
df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]
df1['Total']=df1['Cost per']*df1['Units Sold']
print(df1)

#şimdi openpyxl ile birleştireceğiz
rows= dataframe_to_rows(df1, index=False) #index=false because we don't want to paste extra indices when Excel
# already have them
#print(rows) #böyle dersek çalışmaz, openpyxl list gibi çıktı veriyor çünkü, pandas gibi değil, bunun yerine şöyle
# diyoruz:
#for row in rows:
#    print(row) #böyle yapınca tamamı geldi list formatında

#r_idx: row index
for r_idx, row in enumerate(rows,1):
    for c_idx, col in enumerate(row,6): #regions.xlsx'de 6. kolona kadar data var, o yüzden 6'dan başlattık,
# A'dan başlasın istesek 1 yazardık
        ws.cell(row=r_idx, column=c_idx, value=col)
wb.save('combined.xlsx')