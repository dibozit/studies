from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb=Workbook()
ws=wb.active

ws1 = wb.create_sheet('NewSheet')
ws2 = wb.create_sheet('Another',0)

ws.title = 'MySheet'

#print(wb.sheetnames) #çalıştırınca cevap olarak ['Another', 'MySheet', 'NewSheet'] geliyor , another en başta çünkü
# index'ini 0 verdik

#openpyxl ile çalışırken hayata geçirene kadar ws'ler içinde cell'ler oluşmaz, o yüzden high memory occupy etmeden
# çalışır kodun, bu  bir avantaj imiş

wb2 = load_workbook('regions.xlsx') #existing excel file ekliyorum üstünde çalışmak için

new_sheet = wb2.create_sheet('NewSheet')
active_sheet=wb2.active
cell = active_sheet['A1']
#print(cell) #cell'in konum bilgisini verir
#print(cell.value)

active_sheet['A1'] = 0 #A1'i 0 olarak yazdırdık
wb2.save('Modified2.xlsx')