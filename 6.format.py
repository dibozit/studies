from openpyxl.styles import Font,Color, colors, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle
from openpyxl.workbook import Workbook

wb = Workbook()
ws = wb.active

for i in range(1,20):
    ws.append(range(300))

ws.merge_cells("A1:B5")
ws.unmerge_cells("A1:B5")

#merge from B2 to E5 :
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)
#bu merged cell'i çağırmak için:
cell = ws['B2']
#text'i kırmızı ve italik yaptık, boyunu da 20.
cell.font = Font(color=colors.BLUE, size=20, italic=True)
#değer verdik
cell.value = 'Merged Cell'
#text'in bottom-right'da konumlanması için
cell.alignment = Alignment(horizontal='right', vertical='bottom')
#background of the cell to smooth transition from black to white. solid doldurmak isteseydik pattern fill
# kullanacaktık, ama biz geçişli istediğimiz için radient fill kullandık
cell.fill = GradientFill(stop=("000000","FFFFFF"))
#siyahtan beyaa geçsin isityoruz, RRGGBB renklerinden beyaz olanı için tüm renkler 0 olur, beyazda da hepsi en
# fazlasında yani F'de olur
wb.save('text.xlsx')

highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
#siyah kalın border yapmak için:
bd = Side(style='thick', color='000000')
highlight.border = Border(left=bd, top=bd, right=bd, bottom=bd)
#filling with solid yellow color:
highlight.fill = PatternFill('solid', fgColor='FFFF00')

#apply this style to every cell diagonally starting at the H column:

count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
    col[count].style = highlight
    count= count+1
wb.save('highlight.xlsx')

