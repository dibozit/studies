#openpyxl > visualization konusunda güçlü, small ve medium sized datasetlerle çalışılabilir, slow with large datasets
# pandas > lacks visualisation ama large datasetlerle çok iyi çalışır çünkü excel'i komut çalışana kadar live'a
# geçirmiyor, fast analysis'e müsait
#beraber çalışabiliyorlar, bu sebeple yapiştur

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

#önce dataframe'leri yaratıryorum. hali hazırda bu exceller mevcuttu, çalışma ortamıma attım. hepsinde aynı kolonlar
# mevcut. shifts.xlsx de iki tane sheet var, o yüzden taratırken sheet'i de specify etmem gerekti.
df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx')

#concatenation function will combine all the data frames as long as they have the same header columns. to use the
#concat function, we'll create a new dataframe and pass in all the dataframes as a list.
df_all = pd.concat([df_1, df_2, df_3], sort=False)
#false'u columnlar oldukları yerde kalsınlar diye koyduk
print(df_all) #böylece hepsini getirdik, ama baştaki indexler birden fazla oldu, 50yi getirip deneyelim mesela
print(df_all.loc[50]) #üç tane geldi
print(df_all.groupby(['Shift']).mean()['Units Sold']) #1. 2. ve 3. shift için units sold'un mean'lerini hesaplattık

to_excel = df_all.to_excel('all_shifts.xlsx', index=None)

#şimdi G kolonuna bir de total ekleyelim, bunun için openpyxl lazım, yukarı ekledim

wb= load_workbook('all_shifts.xlsx')
ws= wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)
bd = Side(style='thin', color='000000')
total_col.border = Border(left=bd, top=bd, right=bd, bottom=bd)
total_col.value = 'Total'

e_col, f_col = ['E','F']
#liste 299' kadar gidiyor, bu yüzden range'i 300'e kadar ayarlayacağız, first raw'u da istemiyoruz, çünkü o header value
for row in range(2,300):
    result_cell = 'G{}'.format(row)
    e_value = ws[e_col+str(row)].value
    f_value = ws[f_col+str(row)].value
    ws[result_cell] = e_value*f_value

wb.save('totaled.xlsx')